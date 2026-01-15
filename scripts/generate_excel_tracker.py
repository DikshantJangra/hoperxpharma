#!/usr/bin/env python3
"""
Master Feature & Verification Excel System Generator
Generates a comprehensive 7-sheet Excel file for tracking features, testing, bugs, fixes, and releases.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def create_master_tracker():
    """Create the Master Feature & Verification Excel System"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define color scheme
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ======================
    # 1. REFERENCE DATA SHEET (Create first for validation)
    # ======================
    ref_sheet = wb.create_sheet("Reference Data", 0)
    
    reference_data = {
        'Modules': ['Auth', 'Inventory', 'POS', 'Billing', 'Messages', 'Dashboard', 'Reports', 'Settings', 'Admin'],
        'Feature Status': ['Planned', 'In Development', 'Built', 'Deprecated', 'Removed'],
        'Criticality': ['Low', 'Medium', 'High', 'Blocking'],
        'Verification Status': ['✅ Pass', '❌ Fail', '⏸ Partial', '— Not Tested'],
        'Bug Severity': ['Critical', 'High', 'Medium', 'Low'],
        'Bug Priority': ['P0', 'P1', 'P2', 'P3'],
        'Bug Status': ['Open', 'In Progress', 'Fixed', 'Verified', 'Closed', 'Wont Fix', 'Duplicate'],
        'Fix Type': ['UI', 'Logic', 'Performance', 'Security', 'Data', 'Configuration'],
        'Test Type': ['Unit', 'Integration', 'E2E', 'Manual', 'Automated'],
        'Regression Risk': ['High', 'Medium', 'Low'],
        'Yes/No': ['Yes', 'No', 'N/A'],
        'Team Members': ['Dikshant', 'QA Team', 'Product Team', 'Backend Team', 'Frontend Team']
    }
    
    col_idx = 1
    for category, values in reference_data.items():
        ref_sheet.cell(1, col_idx, category).font = HEADER_FONT
        ref_sheet.cell(1, col_idx).fill = HEADER_FILL
        ref_sheet.cell(1, col_idx).alignment = Alignment(horizontal='center')
        
        for row_idx, value in enumerate(values, start=2):
            ref_sheet.cell(row_idx, col_idx, value)
        
        # Define named ranges for data validation
        col_letter = get_column_letter(col_idx)
        range_name = category.replace(' ', '_').replace('/', '_')
        end_row = len(values) + 1
        
        # Create named range using openpyxl's defined_names
        from openpyxl.workbook.defined_name import DefinedName
        defn = DefinedName(range_name, attr_text=f"'Reference Data'!${col_letter}$2:${col_letter}${end_row}")
        wb.defined_names[range_name] = defn
        
        ref_sheet.column_dimensions[col_letter].width = 20
        col_idx += 1
    
    # ======================
    # 2. FEATURE MASTER SHEET
    # ======================
    master_sheet = wb.create_sheet("Feature Master", 1)
    
    master_headers = [
        'Feature ID', 'Feature Name', 'Module', 'Sub-Module', 'Description',
        'Status', 'Criticality', 'Owner', 'Target Release',
        'Date Added', 'Date Completed', 'Notes'
    ]
    
    for col_idx, header in enumerate(master_headers, start=1):
        cell = master_sheet.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Set column widths
    widths = [12, 30, 15, 20, 40, 15, 12, 15, 15, 12, 12, 30]
    for col_idx, width in enumerate(widths, start=1):
        master_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add data validation
    dv_module = DataValidation(type="list", formula1="Modules", allow_blank=False)
    master_sheet.add_data_validation(dv_module)
    dv_module.add(f'C2:C1000')
    
    dv_status = DataValidation(type="list", formula1="Feature_Status", allow_blank=False)
    master_sheet.add_data_validation(dv_status)
    dv_status.add(f'F2:F1000')
    
    dv_criticality = DataValidation(type="list", formula1="Criticality", allow_blank=False)
    master_sheet.add_data_validation(dv_criticality)
    dv_criticality.add(f'G2:G1000')
    
    dv_owner = DataValidation(type="list", formula1="Team_Members", allow_blank=True)
    master_sheet.add_data_validation(dv_owner)
    dv_owner.add(f'H2:H1000')
    
    # Freeze panes
    master_sheet.freeze_panes = 'B2'
    
    # ======================
    # 3. VERIFICATION CHECKLIST SHEET
    # ======================
    verify_sheet = wb.create_sheet("Verification Checklist", 2)
    
    verify_headers = [
        'Verification ID', 'Feature ID', 'Test Scenario', 'Expected Behavior', 'Test Type',
        'Dev Verified', 'Dev By', 'Dev Date', 'Dev Notes',
        'Tester Verified', 'Tester By', 'Tester Date', 'Tester Notes',
        'Final Status', 'Blocking Issue'
    ]
    
    for col_idx, header in enumerate(verify_headers, start=1):
        cell = verify_sheet.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    
    # Set column widths
    v_widths = [15, 12, 35, 35, 12, 12, 15, 12, 25, 12, 15, 12, 25, 15, 12]
    for col_idx, width in enumerate(v_widths, start=1):
        verify_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add data validation
    dv_test_type = DataValidation(type="list", formula1="Test_Type", allow_blank=False)
    verify_sheet.add_data_validation(dv_test_type)
    dv_test_type.add(f'E2:E1000')
    
    dv_dev_verify = DataValidation(type="list", formula1="Verification_Status", allow_blank=False)
    verify_sheet.add_data_validation(dv_dev_verify)
    dv_dev_verify.add(f'F2:F1000')
    
    dv_tester_verify = DataValidation(type="list", formula1="Verification_Status", allow_blank=False)
    verify_sheet.add_data_validation(dv_tester_verify)
    dv_tester_verify.add(f'J2:J1000')
    
    dv_dev_by = DataValidation(type="list", formula1="Team_Members", allow_blank=True)
    verify_sheet.add_data_validation(dv_dev_by)
    dv_dev_by.add(f'G2:G1000')
    
    dv_tester_by = DataValidation(type="list", formula1="Team_Members", allow_blank=True)
    verify_sheet.add_data_validation(dv_tester_by)
    dv_tester_by.add(f'K2:K1000')
    
    # Add Final Status formula to row 2 (users can copy down)
    verify_sheet['N2'] = '=IF(AND(F2="✅ Pass", J2="✅ Pass"), "✅ Verified", IF(OR(F2="❌ Fail", J2="❌ Fail"), "❌ Failed", IF(OR(F2="⏸ Partial", J2="⏸ Partial"), "⚠️ Partial", "— Pending")))'
    
    verify_sheet.freeze_panes = 'B2'
    
    # ======================
    # 4. BUG TRACKER SHEET
    # ======================
    bug_sheet = wb.create_sheet("Bug Tracker", 3)
    
    bug_headers = [
        'Bug ID', 'Related Feature ID(s)', 'Title', 'Description', 'Severity', 'Priority',
        'Steps to Reproduce', 'Found By', 'Reported By', 'Date Reported',
        'Status', 'Assigned To', 'Fix Reference', 'Date Fixed',
        'Verified By', 'Date Verified', 'Regression Risk', 'Root Cause'
    ]
    
    for col_idx, header in enumerate(bug_headers, start=1):
        cell = bug_sheet.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    
    # Set column widths
    b_widths = [10, 20, 30, 35, 10, 8, 35, 12, 15, 12, 12, 15, 12, 12, 15, 12, 12, 30]
    for col_idx, width in enumerate(b_widths, start=1):
        bug_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add data validation
    dv_severity = DataValidation(type="list", formula1="Bug_Severity", allow_blank=False)
    bug_sheet.add_data_validation(dv_severity)
    dv_severity.add(f'E2:E1000')
    
    dv_priority = DataValidation(type="list", formula1="Bug_Priority", allow_blank=False)
    bug_sheet.add_data_validation(dv_priority)
    dv_priority.add(f'F2:F1000')
    
    dv_found = DataValidation(type="list", formula1="'Reference Data'!L$2:L$5", allow_blank=False)
    dv_found.prompt = "Who discovered this bug?"
    bug_sheet.add_data_validation(dv_found)
    dv_found.add(f'H2:H1000')
    
    dv_bug_status = DataValidation(type="list", formula1="Bug_Status", allow_blank=False)
    bug_sheet.add_data_validation(dv_bug_status)
    dv_bug_status.add(f'K2:K1000')
    
    dv_assigned = DataValidation(type="list", formula1="Team_Members", allow_blank=True)
    bug_sheet.add_data_validation(dv_assigned)
    dv_assigned.add(f'L2:L1000')
    
    dv_regression = DataValidation(type="list", formula1="Regression_Risk", allow_blank=False)
    bug_sheet.add_data_validation(dv_regression)
    dv_regression.add(f'Q2:Q1000')
    
    bug_sheet.freeze_panes = 'B2'
    
    # ======================
    # 5. FIX LOG SHEET
    # ======================
    fix_sheet = wb.create_sheet("Fix Log", 4)
    
    fix_headers = [
        'Fix ID', 'Related Bug ID(s)', 'Feature Impacted', 'Type of Fix', 'Title',
        'Description', 'Files Changed', 'Developer', 'Date Fixed',
        'Verified By', 'Date Verified', 'Release Version', 'Commit/PR Reference',
        'Regression Tests Added'
    ]
    
    for col_idx, header in enumerate(fix_headers, start=1):
        cell = fix_sheet.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    
    # Set column widths
    f_widths = [10, 20, 15, 15, 30, 35, 30, 15, 12, 15, 12, 12, 20, 20]
    for col_idx, width in enumerate(f_widths, start=1):
        fix_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add data validation
    dv_fix_type = DataValidation(type="list", formula1="Fix_Type", allow_blank=False)
    fix_sheet.add_data_validation(dv_fix_type)
    dv_fix_type.add(f'D2:D1000')
    
    dv_developer = DataValidation(type="list", formula1="Team_Members", allow_blank=False)
    fix_sheet.add_data_validation(dv_developer)
    dv_developer.add(f'H2:H1000')
    
    dv_regression_test = DataValidation(type="list", formula1="Yes_No", allow_blank=False)
    fix_sheet.add_data_validation(dv_regression_test)
    dv_regression_test.add(f'N2:N1000')
    
    fix_sheet.freeze_panes = 'B2'
    
    # ======================
    # 6. REGRESSION MATRIX SHEET
    # ======================
    reg_sheet = wb.create_sheet("Regression Matrix", 5)
    
    reg_headers = [
        'Feature ID', 'Feature Name', 'Trigger Condition', 'Dependency Features',
        'Regression Risk', 'Last Regression Test', 'Tested By', 'Result', 'Notes', 'Next Test Due'
    ]
    
    for col_idx, header in enumerate(reg_headers, start=1):
        cell = reg_sheet.cell(1, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    
    # Set column widths
    r_widths = [12, 30, 35, 25, 12, 15, 15, 12, 30, 12]
    for col_idx, width in enumerate(r_widths, start=1):
        reg_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Add data validation
    dv_reg_risk = DataValidation(type="list", formula1="Regression_Risk", allow_blank=False)
    reg_sheet.add_data_validation(dv_reg_risk)
    dv_reg_risk.add(f'E2:E1000')
    
    dv_reg_tested = DataValidation(type="list", formula1="Team_Members", allow_blank=True)
    reg_sheet.add_data_validation(dv_reg_tested)
    dv_reg_tested.add(f'G2:G1000')
    
    dv_result = DataValidation(type="list", formula1="'Reference Data'!D$2:D$5", allow_blank=True)
    dv_result.prompt = "Test result"
    reg_sheet.add_data_validation(dv_result)
    dv_result.add(f'H2:H1000')
    
    # Add VLOOKUP for Feature Name (lookup from Feature Master)
    reg_sheet['B2'] = '=IFERROR(VLOOKUP(A2,\'Feature Master\'!A:B,2,FALSE),"")'
    
    reg_sheet.freeze_panes = 'B2'
    
    # ======================
    # 7. PROGRESS DASHBOARD SHEET
    # ======================
    dash_sheet = wb.create_sheet("Progress Dashboard", 6)
    
    # Title
    dash_sheet.merge_cells('A1:F1')
    title_cell = dash_sheet['A1']
    title_cell.value = "🎯 MASTER FEATURE & VERIFICATION DASHBOARD"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    dash_sheet.row_dimensions[1].height = 30
    
    # Section A: Feature Completion Metrics
    row = 3
    dash_sheet.merge_cells(f'A{row}:F{row}')
    section_cell = dash_sheet[f'A{row}']
    section_cell.value = "📊 FEATURE COMPLETION METRICS"
    section_cell.font = Font(size=12, bold=True)
    section_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    section_cell.font = Font(color="FFFFFF", bold=True)
    
    row += 1
    metrics = [
        ('Total Features', '=COUNTIFS(\'Feature Master\'!F:F,"<>Removed")'),
        ('Features Built', '=COUNTIF(\'Feature Master\'!F:F,"Built")'),
        ('Build Completion %', '=IF(B5>0,ROUND(B6/B5*100,1),0)&"%"'),
        ('Features Verified', '=COUNTIF(\'Verification Checklist\'!N:N,"✅ Verified")'),
        ('Verification %', '=IF(B6>0,ROUND(B8/B6*100,1),0)&"%"'),
    ]
    
    for metric, formula in metrics:
        dash_sheet[f'A{row}'] = metric
        dash_sheet[f'A{row}'].font = Font(bold=True)
        dash_sheet[f'B{row}'] = formula
        dash_sheet[f'B{row}'].font = Font(size=12)
        row += 1
    
    # Section B: Bug Health
    row += 1
    dash_sheet.merge_cells(f'A{row}:F{row}')
    section_cell = dash_sheet[f'A{row}']
    section_cell.value = "🐛 BUG HEALTH"
    section_cell.font = Font(size=12, bold=True)
    section_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    section_cell.font = Font(color="FFFFFF", bold=True)
    
    row += 1
    bug_metrics = [
        ('Open Bugs', '=COUNTIF(\'Bug Tracker\'!K:K,"Open")'),
        ('Critical/High Bugs', '=COUNTIFS(\'Bug Tracker\'!E:E,"Critical",\'Bug Tracker\'!K:K,"<>Closed")+COUNTIFS(\'Bug Tracker\'!E:E,"High",\'Bug Tracker\'!K:K,"<>Closed")'),
        ('Blocking Bugs', '=COUNTIFS(\'Bug Tracker\'!F:F,"P0",\'Bug Tracker\'!K:K,"<>Closed")'),
        ('Average Bug Age (days)', '=IF(COUNTIF(\'Bug Tracker\'!K:K,"Open")>0,ROUND(AVERAGE(IF(\'Bug Tracker\'!K:K="Open",TODAY()-\'Bug Tracker\'!J:J,"")),0),"0")'),
    ]
    
    for metric, formula in bug_metrics:
        dash_sheet[f'A{row}'] = metric
        dash_sheet[f'A{row}'].font = Font(bold=True)
        dash_sheet[f'B{row}'] = formula
        dash_sheet[f'B{row}'].font = Font(size=12)
        row += 1
    
    # Section C: Release Readiness
    row += 1
    dash_sheet.merge_cells(f'A{row}:F{row}')
    section_cell = dash_sheet[f'A{row}']
    section_cell.value = "🚀 RELEASE READINESS"
    section_cell.font = Font(size=12, bold=True)
    section_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    section_cell.font = Font(color="FFFFFF", bold=True)
    
    row += 1
    readiness_criteria = [
        ('All Critical Features Verified', '=IF(COUNTIFS(\'Feature Master\'!G:G,"Blocking",\'Feature Master\'!F:F,"Built")=COUNTIFS(\'Verification Checklist\'!N:N,"✅ Verified",\'Feature Master\'!G:G,"Blocking"),"✅","❌")'),
        ('No Blocking Bugs', '=IF(B13=0,"✅","❌")'),
        ('Verification Coverage ≥ 95%', '=IF(VALUE(LEFT(B9,LEN(B9)-1))>=95,"✅","⚠️")'),
        ('High Severity Bugs ≤ 2', '=IF(B12<=2,"✅","⚠️")'),
    ]
    
    for criteria, formula in readiness_criteria:
        dash_sheet[f'A{row}'] = criteria
        dash_sheet[f'A{row}'].font = Font(bold=True)
        dash_sheet[f'B{row}'] = formula
        dash_sheet[f'B{row}'].font = Font(size=14)
        row += 1
    
    # Overall Release Status
    row += 1
    dash_sheet[f'A{row}'] = "OVERALL RELEASE STATUS"
    dash_sheet[f'A{row}'].font = Font(bold=True, size=14)
    dash_sheet[f'B{row}'] = '=IF(COUNTIF(B18:B21,"❌")>0,"🔴 BLOCKED",IF(COUNTIF(B18:B21,"⚠️")>0,"🟡 AT RISK","🟢 READY"))'
    dash_sheet[f'B{row}'].font = Font(size=16, bold=True)
    
    # Set column widths
    dash_sheet.column_dimensions['A'].width = 35
    dash_sheet.column_dimensions['B'].width = 20
    
    # Instructions section
    row += 2
    dash_sheet.merge_cells(f'A{row}:F{row}')
    inst_cell = dash_sheet[f'A{row}']
    inst_cell.value = "📖 HOW TO USE THIS DASHBOARD"
    inst_cell.font = Font(size=11, bold=True)
    inst_cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
    inst_cell.font = Font(color="FFFFFF", bold=True)
    
    row += 1
    instructions = [
        "1. This dashboard auto-updates based on data in other sheets",
        "2. Green (✅/🟢) = Ready | Yellow (⚠️/🟡) = At Risk | Red (❌/🔴) = Blocked",
        "3. Review blocking items before any release",
        "4. All formulas pull from Feature Master, Verification Checklist, and Bug Tracker",
        "5. Update this dashboard by refreshing (Ctrl+Alt+F9 or Cmd+Option+F9)"
    ]
    
    for instruction in instructions:
        dash_sheet[f'A{row}'] = instruction
        dash_sheet[f'A{row}'].font = Font(size=9, italic=True)
        row += 1
    
    # ======================
    # FINAL FORMATTING
    # ======================
    
    # Set print settings for all sheets
    for sheet in wb.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
    
    # Save the workbook
    output_path = '/Users/dikshantjangra/Desktop/hoperxpharma/Master_Feature_Verification_System.xlsx'
    wb.save(output_path)
    print(f"✅ Excel file created successfully: {output_path}")
    print(f"📊 Total sheets created: {len(wb.worksheets)}")
    print(f"🎯 System ready for use!")
    
    return output_path

if __name__ == "__main__":
    create_master_tracker()
